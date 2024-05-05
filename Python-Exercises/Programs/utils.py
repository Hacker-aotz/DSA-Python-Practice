# Finding Max Value
from openpyxl.chart import BarChart, Reference
import openpyxl as xl


def find_max(numbers):
    maximum = numbers[0]
    for number in numbers:
        if number > maximum:
            maximum = number
    return maximum


# Emoji Funtion
def emoji_converter(message):
    words = message.split(" ")
    emoji = {
        ":)": "🙂",
        ":(": "🙁"
    }
    output = ""
    for word in words:
        output += emoji.get(word, word) + " "
    return output


# Processing Workbook


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * .9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, "e2")

    wb.save(filename)


# Singly Linked List:
class Node:
    def __init__(self, value=None):
        self.value = value
        self.next = None


class SLinkedList:
    def __init__(self):
        self.head = None
        self.tail = None

    def __iter__(self):
        node = self.head
        while node:
            yield node
            node = node.next

    # Insertion of element:

    def insertSLL(self, value, location):
        newNode = Node(value)
        if self.head is None:
            self.head = newNode
            self.tail = newNode
        else:
            if location == 0:
                newNode.next = self.head
                self.head = newNode
            elif location == -1:
                newNode.next = None
                self.tail.next = newNode
                self.tail = newNode
            else:
                tempNode = self.head
                index = 0
                while index < location - 1:
                    tempNode = tempNode.next
                    index += 1
                nextNode = tempNode.next
                tempNode.next = newNode
                newNode.next = nextNode
                if tempNode == self.tail:
                    self.tail = newNode

    # Traverse Singly Linked List:

    def traverseSLL(self):
        if self.head is None:
            print("Element does not exist")
        else:
            node = self.head
            while node is not None:
                print(node.value)
                node = node.next

    # Search for a node in Singly Linked List:

    def searchSLL(self, nodeValue):
        if self.head is None:
            return "The Singly linked list doesn't exist."
        else:
            node = self.head
            while node is not None:
                if node.value == nodeValue:
                    return node.value
                node = node.next
            return "Value doesn't exist in the list."

    #  Delete a node from Singly Linked List

    def deleteNode(self, location):
        if self.head is None:
            print("The Singly linked list doesn't exist.")
        else:
            if location == 0:
                if self.head == self.tail:
                    self.head = None
                    self.tail = None
                else:
                    self.head = self.head.next
            elif location == -1:
                if self.head == self.tail:
                    self.head = None
                    self.tail = None
                else:
                    node = self.head
                    while node is not None:
                        if node.next == self.tail:
                            break
                        node = node.next
                    node.next = None
                    self.tail = node
            else:
                tempNode = self.head
                index = 0
                while index < location - 1:
                    tempNode = tempNode.next
                    index += 1
                nextNode = tempNode.next
                tempNode.next = nextNode.next

    # Delete entire SLL

    def deleteESLL(self):
        if self.head == None:
            print("The SLL doesn't exist.")
        else:
            self.head = None
            self.tail = None


singlyLinkedList = SLinkedList()
singlyLinkedList.insertSLL(1, 1)
singlyLinkedList.insertSLL(2, 1)
singlyLinkedList.insertSLL(5, 1)
singlyLinkedList.insertSLL(4, 1)

singlyLinkedList.insertSLL(10, 0)

singlyLinkedList.insertSLL(9, -1)

singlyLinkedList.insertSLL(6, 4)

print([node.value for node in singlyLinkedList])

singlyLinkedList.deleteESLL()

print([node.value for node in singlyLinkedList])
